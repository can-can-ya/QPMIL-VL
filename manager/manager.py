import torch
import torch.nn as nn
import os.path as osp
import wandb
from types import SimpleNamespace
from torch.optim import Adam, lr_scheduler
from torch import Tensor
from torch.nn.functional import softmax
import pandas as pd
import re
import ast
import openpyxl

from utils import get_device, BinClf_Evaluator, MultiClf_Evaluator, get_loss_function, get_init_key_frequency, mkdir_if_missing, get_current_ensemble_classes, rename_keys, check_feature_distribution, plot_key_matching_heatmap, get_current_eval_dataloader
from models.conch import create_model_from_pretrained
from models import freeze_weight, set_tunable_v, QPMIL_VL, EarlyStopping
from dataset import get_data_loaders, get_sids


class Manager(object):
    def __init__(self, cfg):
        self.cfg = cfg

        self.device = get_device(cfg['cuda_id'])

        if cfg['base_model_arch'] == 'CONCH':
            self.base_model, _ = create_model_from_pretrained("conch_ViT-B-16", checkpoint_path=cfg['conch_ckpt_path'], device=self.device)
            self.base_model.eval()
            self.base_model.dtype = self.base_model.logit_scale.dtype
            self.embedding_dim = self.base_model.text.ln_final.weight.shape[0] # D_e: embedding dimension of E_txt
            self.feature_dim = self.base_model.visual.proj_contrast.shape[1] # D_f: output feature dimension of encoder E_img and E_txt
        else:
            raise NotImplementedError("Please specify a valid architecture.")
        freeze_weight(self.base_model, cfg['base_model_arch'])
        self.dtype = self.base_model.dtype

        self.writer = None

        self.model = None
        self.optimizer = None
        self.lr_scheduler = None
        self.early_stop = None

        self.b_evaluator = BinClf_Evaluator()
        self.m_evaluator = MultiClf_Evaluator()

        self.loss_function = get_loss_function(cfg['loss_function'])
        print('[setup] loss function:', cfg['loss_function'])

        self.current_dataset = None
        self.current_save_result_dir = None
        self.current_ensemble_classes = None

        self.data_loader = get_data_loaders(cfg)
        self.sids = get_sids(self.data_loader)

        # Parameters to be trained: Prototype Pool (key, prompt)
        if cfg['opt_name'] == 'adam':
            self.key = nn.ParameterList([nn.Parameter(0.02 * torch.randn(1, self.feature_dim, dtype=self.dtype, device=self.device)) for _ in range(cfg['pool_size'])])
            self.prompt = nn.ParameterList([nn.Parameter(0.02 * torch.randn(1, cfg['prompt_length'], self.embedding_dim, dtype=self.dtype, device=self.device)) for _ in range(cfg['pool_size'])])
            print('[setup] Parameters to be trained:\nPrototype Pool\nkey:\n{}\nprompt:\n{}'.format(self.key, self.prompt))
        else:
            raise NotImplementedError('Invalid optimizer')

        # Parameters to be trained: Tunable Vector
        self.tunable_v = nn.ParameterList([nn.Parameter(torch.zeros(class_num, self.feature_dim, dtype=self.dtype, device=self.device)) for class_num in cfg['dataset_subtype_num']])
        print('Tunable Vector\n{}'.format(self.tunable_v))

        self.train_key_frequency = get_init_key_frequency(cfg)
        self.val_key_frequency = get_init_key_frequency(cfg)

        self.sum_test_acc = 0
        self.test_acc = []

    def incre_train(self):
        for dataset_name in self.cfg['dataset_names']:
            self.cfg['task_num'] += 1
            print('[train] Task: {}, Dataset: {}'.format(self.cfg['task_num'], dataset_name))

            self.current_dataset = dataset_name
            self.current_save_result_dir = osp.join(self.cfg['save_result_dir'], dataset_name+'-task'+str(self.cfg['task_num']))
            mkdir_if_missing(self.current_save_result_dir)
            self.current_ensemble_classes = get_current_ensemble_classes(self.cfg, self.current_dataset)
            print('[infor] Current dataset: {}'.format(self.current_dataset))
            print('[infor] Current save result dir: {}'.format(self.current_save_result_dir))
            print('[infor] Current ensemble classes: {}'.format(self.current_ensemble_classes))

            # wandb writer
            project_name = 'QPMIL-VL_' + self.cfg['save_result_dir'].split('/')[1] + '_' + self.cfg['save_result_dir'].split('/')[2]
            run_name = self.cfg['save_result_dir'].split('/')[-1] + '-' + dataset_name + '-task' + str(self.cfg['task_num'])
            self.writer = wandb.init(project=project_name, name=run_name, dir=self.current_save_result_dir, config=self.cfg, reinit=True)
            print('[setup] Wandb writer construction completed')

            # model
            set_tunable_v(self.tunable_v, self.cfg['task_num']) # Only the tunable vectors of the current dataset classes are trained.
            self.model = QPMIL_VL(self.cfg, self.base_model, self.device, self.key, self.prompt, self.tunable_v, self.current_ensemble_classes, self.train_key_frequency)
            print('[setup] Model construction completed')

            # optimizer
            if self.cfg['opt_name'] == 'adam':
                cfg_optimizer = SimpleNamespace(opt=self.cfg['opt_name'], weight_decay=self.cfg['adam_weight_decay'],
                                                lr=self.cfg['adam_lr'], opt_eps=self.cfg['adam_eps'], opt_betas=None,
                                                momentum=None)
                self.optimizer = Adam(self.model.parameters(), lr=cfg_optimizer.lr,
                                      weight_decay=cfg_optimizer.weight_decay, eps=cfg_optimizer.opt_eps)
            else:
                raise NotImplementedError('Invalid optimizer')
            print("[setup] optimizer:", cfg_optimizer)

            # lr_scheduler
            cfg_lr_scheduler = SimpleNamespace(name=self.cfg['lrs_name'], mode=self.cfg['lrs_mode'],
                                               factor=self.cfg['lrs_factor'],
                                               patience=self.cfg['lrs_patience'][self.cfg['task_num'] - 1],
                                               threshold=self.cfg['lrs_threshold'],
                                               threshold_mode=self.cfg['lrs_threshold_mode'],
                                               verbose=self.cfg['lrs_verbose'])
            if cfg_lr_scheduler.name == 'ReduceLROnPlateau':
                self.lr_scheduler = lr_scheduler.ReduceLROnPlateau(self.optimizer, mode=cfg_lr_scheduler.mode,
                                                                   factor=cfg_lr_scheduler.factor,
                                                                   patience=cfg_lr_scheduler.patience,
                                                                   threshold=cfg_lr_scheduler.threshold,
                                                                   threshold_mode=cfg_lr_scheduler.threshold_mode,
                                                                   verbose=cfg_lr_scheduler.verbose)
            else:
                raise NotImplementedError('Invalid lr scheduler')
            print("[setup] lr_scheduler:", cfg_lr_scheduler)

            # early_stop
            self.early_stop = EarlyStopping(warmup=self.cfg['es_warmup'],
                                            patience=self.cfg['es_patience'][self.cfg['task_num'] - 1],
                                            verbose=self.cfg['es_verbose'], threshold=self.cfg['es_threshold'])
            print(f"[setup] early_stop: warmup={self.cfg['es_warmup']}, patience={self.cfg['es_patience'][self.cfg['task_num'] - 1]}, verbose={self.cfg['es_verbose']}, threshold={self.cfg['es_threshold']}")

            # run training
            self._run_training()

            # load best checkpoint
            if self.cfg['load_best_ckpt']:
                print('[infor] loading best checkpoint...')
                best_ckpt_path = osp.join(self.current_save_result_dir, 'model_ckpts', 'best.pth')
                checkpoint = torch.load(best_ckpt_path, map_location=self.device)

                if self.cfg['opt_name'] == 'adam':
                    for key, value in checkpoint.items():
                        if 'key' in key or 'prompt' in key or 'tunable_v' in key:
                            self.model.state_dict()[key].copy_(value)
                else:
                    raise NotImplementedError('Invalid optimizer')

            # evaluate at val/test dataset
            eval_result_dir = osp.join(self.current_save_result_dir, 'eval_results')
            mkdir_if_missing(eval_result_dir)
            current_eval_dataloader = get_current_eval_dataloader(self.data_loader, self.current_dataset)
            self._eval_all(current_eval_dataloader, eval_result_dir)
            print('[eval at val/test dataset] Path where the evaluation result is saved:', eval_result_dir)

            # calculate masked metrics at val/test dataset
            masked_eval_result_dir = eval_result_dir
            self._eval_masked_metrics(current_eval_dataloader, masked_eval_result_dir)
            print('[cal masked metrics at val/test dataset] Path where the evaluation result is saved:', masked_eval_result_dir)

        average_test_acc = self.sum_test_acc / len(self.cfg['dataset_names'])
        print('[result] Average test acc: {:.6f}'.format(average_test_acc))
        wandb.log({'result/avg_test_acc': average_test_acc})

        if self.cfg['data_split_seed'] == self.cfg['total_fold']:
            metrics_dir = osp.join(osp.dirname(osp.dirname(self.current_save_result_dir)), 'metrics')
            mkdir_if_missing(metrics_dir)

            for fold_num in range(1, self.cfg['total_fold'] + 1):
                with open(osp.join(metrics_dir, 'test_acc.txt'), 'a') as f:
                    read_path = osp.join(re.sub(r'_\d+$', '_' + str(fold_num), osp.dirname(self.current_save_result_dir)), 'metrics/test_acc.txt')
                    with open(read_path, 'r') as file_read:
                        f.write(file_read.read() + '\n')

                with open(osp.join(metrics_dir, 'val_acc.txt'), 'a') as f:
                    read_path = osp.join(re.sub(r'_\d+$', '_' + str(fold_num), osp.dirname(self.current_save_result_dir)), 'metrics/val_acc.txt')
                    with open(read_path, 'r') as file_read:
                        f.write(file_read.read() + '\n')

                with open(osp.join(metrics_dir, 'test_mask_acc.txt'), 'a') as f:
                    read_path = osp.join(re.sub(r'_\d+$', '_' + str(fold_num), osp.dirname(self.current_save_result_dir)), 'metrics/test_mask_acc.txt')
                    with open(read_path, 'r') as file_read:
                        f.write(file_read.read() + '\n')

                with open(osp.join(metrics_dir, 'test_mask_auc.txt'), 'a') as f:
                    read_path = osp.join(re.sub(r'_\d+$', '_' + str(fold_num), osp.dirname(self.current_save_result_dir)), 'metrics/test_mask_auc.txt')
                    with open(read_path, 'r') as file_read:
                        f.write(file_read.read() + '\n')

            result_metric_list = []
            for fold_num in range(1, self.cfg['total_fold'] + 1):
                read_path = osp.join(re.sub(r'_\d+$', '_' + str(fold_num), osp.dirname(self.current_save_result_dir)), 'metrics/test_acc.txt')
                with open(read_path, 'r') as file_read:
                    result_metric_list.extend(ast.literal_eval(file_read.read()))
            for fold_num in range(1, self.cfg['total_fold'] + 1):
                read_path = osp.join(re.sub(r'_\d+$', '_' + str(fold_num), osp.dirname(self.current_save_result_dir)), 'metrics/test_mask_acc.txt')
                with open(read_path, 'r') as file_read:
                    result_metric_list.append(ast.literal_eval(file_read.read()))
            for fold_num in range(1, self.cfg['total_fold'] + 1):
                read_path = osp.join(re.sub(r'_\d+$', '_' + str(fold_num), osp.dirname(self.current_save_result_dir)), 'metrics/test_mask_auc.txt')
                with open(read_path, 'r') as file_read:
                    result_metric_list.append(ast.literal_eval(file_read.read()))
            for fold_num in range(1, self.cfg['total_fold'] + 1):
                read_path = osp.join(re.sub(r'_\d+$', '_' + str(fold_num), osp.dirname(self.current_save_result_dir)), 'metrics/val_acc.txt')
                with open(read_path, 'r') as file_read:
                    result_metric_list.append(ast.literal_eval(file_read.read()))

            wb = openpyxl.load_workbook(self.cfg['eval_template_path'])
            ws = wb.active
            ws.cell(row=2, column=2, value=osp.dirname(osp.dirname(self.current_save_result_dir)).split('/')[-1])
            for i in range(4):
                ws.cell(row=3, column=i+5, value=self.cfg['dataset_names'][i])
            for row_index, row in enumerate(result_metric_list, start=4):
                for col_index, value in enumerate(row, start=5):
                    ws.cell(row=row_index, column=col_index, value=value)
            wb.save(osp.join(metrics_dir, osp.dirname(osp.dirname(self.current_save_result_dir)).split('/')[-1] + ' result.xlsx'))
            wb.close()

    def _eval_masked_metrics(self, data_loaders, masked_eval_result_dir):
        if self.current_dataset == self.cfg['dataset_names'][-1]:
            test_mask_acc = []
            test_mask_auc = []

        for k, data_loader in data_loaders.items():
            cltor = self.eval_model(self.model, data_loader)

            for k_cltor, v_cltor in cltor.items():
                label_shift = self.cfg['dataset_label_shift'][self.cfg['dataset_names'].index(k.split('/')[0])]
                v_cltor['y_hat'] = v_cltor['y_hat'][:, label_shift:label_shift + self.cfg['dataset_subtype_num'][self.cfg['dataset_names'].index(k.split('/')[0])]] # mask irrelevant logits
                if_binary = v_cltor['y_hat'].shape[1] == 2
                eval_results = self._eval_and_print(v_cltor, name=k + '/masked/' + k_cltor, if_binary=if_binary)

                if self.current_dataset == self.cfg['dataset_names'][-1]:
                    if k.split('/')[-1] == 'test':
                        mask_acc = eval_results['acc@mid'] if if_binary else eval_results['acc']
                        mask_acc = round(mask_acc, 6)
                        test_mask_acc.append(mask_acc)

                        mask_auc = eval_results['auc']
                        mask_auc = round(mask_auc, 6)
                        test_mask_auc.append(mask_auc)

                sids = self._get_unique_sids(k, v_cltor['idx'])
                path_save_pred = masked_eval_result_dir + '/masked-' + k.replace('/', '-') + '.csv'
                self._save_prediction_clf(sids, v_cltor['y'], v_cltor['y_hat'], path_save_pred, binary=if_binary)

        if self.current_dataset == self.cfg['dataset_names'][-1]:
            metrics_dir = osp.join(osp.dirname(self.current_save_result_dir), 'metrics')
            mkdir_if_missing(metrics_dir)

            with open(osp.join(metrics_dir, 'test_mask_acc.txt'), 'w') as f:
                f.write(str(test_mask_acc))

            with open(osp.join(metrics_dir, 'test_mask_auc.txt'), 'w') as f:
                f.write(str(test_mask_auc))

    def _eval_all(self, data_loaders, eval_result_dir):
        current_test_acc = []

        if self.current_dataset == self.cfg['dataset_names'][-1]:
            val_acc = []

        for k, data_loader in data_loaders.items():
            cltor = self.eval_model(self.model, data_loader)

            for k_cltor, v_cltor in cltor.items():
                v_cltor['y'] += self.cfg['dataset_label_shift'][self.cfg['dataset_names'].index(k.split('/')[0])] # label shift
                if_binary = len(self.current_ensemble_classes['count']) == 2
                eval_results = self._eval_and_print(v_cltor, name=k+'/'+k_cltor, if_binary=if_binary)

                if self.current_dataset == self.cfg['dataset_names'][-1] and k.split('/')[-1] == 'test':
                    self.sum_test_acc += eval_results['acc@mid'] if if_binary else eval_results['acc']

                if k.split('/')[-1] == 'test':
                    acc = eval_results['acc@mid'] if if_binary else eval_results['acc']
                    acc = round(acc, 6)
                    current_test_acc.append(acc)

                if self.current_dataset == self.cfg['dataset_names'][-1]:
                    if k.split('/')[-1] == 'val':
                        acc = eval_results['acc@mid'] if if_binary else eval_results['acc']
                        acc = round(acc, 6)
                        val_acc.append(acc)

                sids = self._get_unique_sids(k, v_cltor['idx'])
                path_save_pred = eval_result_dir + '/' + k.replace('/', '-') + '.csv'
                self._save_prediction_clf(sids, v_cltor['y'], v_cltor['y_hat'], path_save_pred, binary=if_binary)

        self.test_acc.append(current_test_acc)

        if self.current_dataset == self.cfg['dataset_names'][-1]:
            metrics_dir = osp.join(osp.dirname(self.current_save_result_dir), 'metrics')
            mkdir_if_missing(metrics_dir)

            with open(osp.join(metrics_dir, 'test_acc.txt'), 'w') as f:
                f.write(str(self.test_acc))

            with open(osp.join(metrics_dir, 'val_acc.txt'), 'w') as f:
                f.write(str(val_acc))

    def _get_unique_sids(self, k, idxs, concat=None):
        dataset_name = k.split('/')[0]
        dataset_split = k.split('/')[-1]
        sids = self.sids[dataset_name][dataset_split]
        idxs = idxs.tolist()
        if concat is None:
            return [sids[i] for i in idxs]
        else:
            return [sids[v] + "-" + str(concat[i].item()) for i, v in enumerate(idxs)]

    def _save_prediction_clf(self, sids, y_true, y_pred, save_path, binary=True, forgetting=False):
        if isinstance(y_true, Tensor):
            y_true = y_true.numpy()
        if isinstance(y_pred, Tensor):
            # pred is logit
            y_pred = softmax(y_pred, dim=1)  # apply softmax to logit outputs (N, num_cls)
            y_pred = y_pred.numpy()

        assert len(sids) == len(y_true)
        assert len(sids) == len(y_pred)

        save_data = {'sids': sids, 'y': y_true}
        cols = ['sids', 'y']
        if binary:
            save_data['y_hat'] = y_pred[:, 1]
            cols.append('y_hat')
        else:
            for i in range(y_pred.shape[-1]):
                _col = 'y_hat_' + str(i)
                save_data[_col] = y_pred[:, i]
                cols.append(_col)

        df = pd.DataFrame(save_data, columns=cols)
        df.to_csv(save_path, index=False)

        wandb_display_name = osp.splitext(osp.basename(save_path))[0].replace('-', '/')+"/dataframe"
        wandb.log({'forgetting_process/'+wandb_display_name if forgetting else wandb_display_name: wandb.Table(dataframe=df)})

    def _run_training(self):
        # iterative training
        last_epoch = -1

        for epoch in range(self.cfg['epochs'][self.cfg['task_num'] - 1]):
            last_epoch = epoch + 1
            print('[train] fold: {}, task: {}, epoch: {}'.format(self.cfg['data_split_seed'], self.cfg['task_num'], last_epoch))
            print('[train] lr={:.8f}'.format(self.optimizer.param_groups[0]['lr']))
            wandb.log({'train/lr/val': self.optimizer.param_groups[0]['lr']})

            # val datasets of previous tasks are also evaluated during training
            # supervise the catastrophic forgetting process
            if self.cfg['check_forgetting_process']:
                for k, v in self.data_loader.items():
                    if k == self.current_dataset:
                        break
                    eval_cltor = self.eval_model(self.model, v['val'])
                    for k_cltor, v_cltor in eval_cltor.items():
                        v_cltor['y'] += self.cfg['dataset_label_shift'][self.cfg['dataset_names'].index(k)]  # label shift
                        if_binary = len(self.current_ensemble_classes['count']) == 2
                        self._eval_and_print(v_cltor, name='forgetting_process/' + k + '/val/' + k_cltor, at_epoch=last_epoch, if_binary=if_binary)

                        sids = self._get_unique_sids(k + '/val', v_cltor['idx'])
                        forgetting_save_dir = osp.join(self.current_save_result_dir, 'forgetting_process')
                        mkdir_if_missing(forgetting_save_dir)
                        forgetting_save_path = osp.join(forgetting_save_dir, k + '-val.csv')
                        self._save_prediction_clf(sids, v_cltor['y'], v_cltor['y_hat'], forgetting_save_path, binary=if_binary, forgetting=True)

            train_cltor = self._train_each_epoch()

            for k_cltor, v_cltor in train_cltor.items():
                self.train_key_frequency[self.current_dataset].append(v_cltor['all_key_frequencies'])
                if_binary = len(self.current_ensemble_classes['count']) == 2
                self._eval_and_print(v_cltor, name='train/'+k_cltor, at_epoch=last_epoch, if_binary=if_binary)

            eval_results = None
            eval_cltor = self.eval_model(self.model, self.data_loader[self.current_dataset]['val'])
            for k_cltor, v_cltor in eval_cltor.items():
                self.val_key_frequency[self.current_dataset].append(v_cltor['all_key_frequencies'])

                v_cltor['y'] += self.cfg['dataset_label_shift'][self.cfg['task_num'] - 1] # label shift
                if_binary = len(self.current_ensemble_classes['count']) == 2
                eval_results = self._eval_and_print(v_cltor, name='val/'+k_cltor, at_epoch=last_epoch, if_binary=if_binary)
                val_loss = eval_results['loss']
                val_acc = eval_results['acc@mid'] if if_binary else eval_results['acc']
                if self.cfg['only_val_loss']:
                    monitor_metric = val_loss
                else:
                    monitor_metric = (val_loss + (1 - val_acc)) / 2

            if last_epoch > self.cfg['lrs_warmup']:
                self.lr_scheduler.step(monitor_metric)
            self.early_stop(last_epoch, monitor_metric)

            if self.early_stop.save_ckpt():
                self._save_model(last_epoch, ckpt_type='best')
                print("[save best model] best model saved at epoch {}".format(last_epoch))

                if self.cfg['check_feature']:
                    save_dir = osp.join(self.current_save_result_dir, 'check_feature')
                    mkdir_if_missing(save_dir)
                    for k_cltor, v_cltor in train_cltor.items():
                        check_feature_distribution(v_cltor['all_feature_dict']['image'], v_cltor['y'], save_dir, type='image', epoch=last_epoch)
                        check_feature_distribution(v_cltor['all_feature_dict']['text'], v_cltor['y'], save_dir, type='text', epoch=last_epoch)
            if self.early_stop.stop():
                break

        self._save_model(last_epoch, ckpt_type='last')
        print("[save last model] last model saved at epoch {}".format(last_epoch))

        if self.cfg['check_key_matching']:
            print('[all key matching frequency on train dataset]', self.train_key_frequency[self.current_dataset])
            print('[all key matching frequency on validation dataset]', self.val_key_frequency[self.current_dataset])
            key_matching_save_dir = osp.join(self.current_save_result_dir, 'key_matching')
            mkdir_if_missing(key_matching_save_dir)
            plot_key_matching_heatmap(self.train_key_frequency[self.current_dataset], key_matching_save_dir, dataset='train')
            plot_key_matching_heatmap(self.val_key_frequency[self.current_dataset], key_matching_save_dir, dataset='val')

    def _save_model(self, epoch, ckpt_type='best'):
        net_ckpt_dict = self._get_state_dict(epoch)
        save_dir = osp.join(self.current_save_result_dir, 'model_ckpts')
        mkdir_if_missing(save_dir)
        torch.save(net_ckpt_dict, save_dir+f'/{ckpt_type}.pth')

    def _get_state_dict(self, epoch=None):
        if self.cfg['opt_name'] == 'adam':
            return_dict = {'epoch': epoch}
            return_dict.update({key: val for key, val in self.model.state_dict().items() if 'key' in key or 'prompt' in key or 'tunable_v' in key})
        else:
            raise NotImplementedError('Invalid optimizer')

        return return_dict

    @torch.no_grad()
    def eval_model(self, model, loader, ckpt_path=None):
        if ckpt_path is not None:
            pass
        model.eval()

        idx_collector, x_collector, y_collector = [], [], []
        all_idx, all_pred, all_gt, all_key_frequencies = [], [], [], torch.zeros(self.cfg['pool_size'], dtype=torch.int)
        i_batch = 0
        eval_every_batch = self.cfg['eval_every_batch']

        for data_idx, data_x, data_y in loader: # [B, 1], [B, N, C], [B, 1]: B = 1
            i_batch += 1

            data_x = data_x.to(self.device)

            x_collector.append(data_x)
            y_collector.append(data_y)
            idx_collector.append(data_idx)

            # in a mini_batch
            if i_batch % eval_every_batch == 0:
                # evaluate network
                cur_pred, key_indices = model(x_collector, eval=True) # [MB, num_cls]
                all_pred.append(cur_pred.detach().cpu())
                all_gt.append(torch.cat(y_collector, dim=0))
                all_idx.append(torch.cat(idx_collector, dim=0))

                key_indices = key_indices.detach().cpu()
                key_frequencies = torch.zeros(self.cfg['pool_size'], dtype=torch.int)
                for index in key_indices.flatten():
                    key_frequencies[index] += 1
                all_key_frequencies += key_frequencies

                # reset mini-batch
                idx_collector, x_collector, y_collector = [], [], []
                torch.cuda.set_device(self.cfg['cuda_id'])
                torch.cuda.empty_cache()

        all_pred = torch.cat(all_pred, dim=0)  # [num_slide, num_cls]
        all_gt = torch.cat(all_gt, dim=0).squeeze(1)  # [num_slide, ]
        all_idx = torch.cat(all_idx, dim=0).squeeze(1)  # [num_slide, ]

        eval_cltor = dict()
        eval_cltor['pred'] = {'y': all_gt, 'y_hat': all_pred, 'idx': all_idx, 'all_key_frequencies': all_key_frequencies}

        return eval_cltor

    def _eval_and_print(self, cltor, name='', at_epoch=None, if_binary=True):
        if if_binary:
            eval_results = self.b_evaluator.compute(cltor, self.cfg['b_metrics_list'], self.cfg['loss_function'])
        else:
            eval_results = self.m_evaluator.compute(cltor, self.cfg['m_metrics_list'], self.cfg['loss_function'])
        wandb_eval_results = rename_keys(eval_results, name, sep='/')

        print("[{}] At epoch {}:".format(name, at_epoch), end=' ')
        print(' '.join(['{}={:.6f},'.format(k, v) for k, v in eval_results.items()]))
        wandb.log(wandb_eval_results)

        return eval_results

    def _train_each_epoch(self):
        self.model.eval()

        idx_collector, x_collector, y_collector = [], [], []
        all_pred, all_gt, all_key_frequencies, all_feature_dict = [], [], torch.zeros(self.cfg['pool_size'], dtype=torch.int), {'image': [], 'text': []}
        i_batch = 0
        train_loader = self.data_loader[self.current_dataset]['train']
        bp_every_batch = self.cfg['bp_every_batch']

        for data_idx, data_x, data_y in train_loader: # [B, 1], [B, N, C], [B, 1]: B = 1
            i_batch += 1

            data_y += self.cfg['dataset_label_shift'][self.cfg['task_num'] - 1] # label shift
            data_x = data_x.to(self.device)
            data_y = data_y.to(self.device)

            x_collector.append(data_x)
            y_collector.append(data_y)
            idx_collector.append(data_idx)

            # in a mini-batch
            if i_batch % bp_every_batch == 0:
                # update network
                cur_pred, key_indices, feature_dict = self._update_network(i_batch, x_collector, y_collector, idx_collector) # [MB, num_cls]
                all_pred.append(cur_pred.detach().cpu())
                all_gt.append(torch.cat(y_collector, dim=0).detach().cpu())

                key_indices = key_indices.detach().cpu()
                key_frequencies = torch.zeros(self.cfg['pool_size'], dtype=torch.int)
                for index in key_indices.flatten():
                    key_frequencies[index] += 1
                all_key_frequencies += key_frequencies
                if self.cfg['print_mb_key_matching']:
                    print('[key matching frequencies]', key_frequencies.tolist())

                all_feature_dict['image'].append(feature_dict['image'].detach().cpu())
                all_feature_dict['text'].append(feature_dict['text'].detach().cpu())

                # reset mini-batch
                idx_collector, x_collector, y_collector = [], [], []
                torch.cuda.set_device(self.cfg['cuda_id'])
                torch.cuda.empty_cache()

        all_pred = torch.cat(all_pred, dim=0)  # [num_slide, num_cls]
        all_gt = torch.cat(all_gt, dim=0).squeeze(1)  # [num_slide, ]
        all_feature_dict['image'] = torch.cat(all_feature_dict['image'], dim=0) # [num_slide, C]
        all_feature_dict['text'] = torch.cat(all_feature_dict['text'], dim=0) # [num_slide, num_cls, C]

        train_cltor = dict()
        train_cltor['pred'] = {'y': all_gt, 'y_hat': all_pred, 'all_key_frequencies': all_key_frequencies, 'all_feature_dict': all_feature_dict}

        return train_cltor

    def _update_network(self, i_batch, xs, ys, idx_num):
        """
        Update network using one mini-batch data
        """
        bag_preds, loss_dict, key_indices, feature_dict = self.model(xs) # [MB, num_cls]

        # zero gradients buffer
        self.optimizer.zero_grad()

        # classification loss
        bag_label = torch.cat(ys, dim=0).squeeze(-1)  # [MB, ]
        clf_loss = self.loss_function(bag_preds, bag_label)
        # total loss
        total_loss = clf_loss + self.cfg['lambda'] * loss_dict['matching_loss'] + self.cfg['beta'] * loss_dict['class_sim_loss']

        print(
            "[training one epoch] {}-th batch: total loss = {:.6f}, classification loss = {:.6f}, matching loss = {:.6f}, class similarity loss = {:.6f}".format(
                i_batch, total_loss.item(), clf_loss.item(), loss_dict['matching_loss'].item(),
                loss_dict['class_sim_loss'].item()))
        wandb_dict = {'train/total loss': total_loss.item(), 'train/classification loss': clf_loss.item(),
                      'train/matching loss': loss_dict['matching_loss'].item(),
                      'train/class similarity loss': loss_dict['class_sim_loss'].item()}
        wandb.log(wandb_dict)

        # backward gradients and update networks
        total_loss.backward()
        if self.cfg['max_norm'] != 'None':
            torch.nn.utils.clip_grad_norm_(self.model.parameters(), self.cfg['max_norm'])
        if self.cfg['opt_name'] == 'adam':
            for param in self.model.parameters():
                if param.grad is not None:
                    if torch.all(param.grad == 0):
                        param.grad = None # Only the matched key and prompt are updated.
        else:
            raise NotImplementedError('Invalid optimizer')
        self.optimizer.step()

        return bag_preds, key_indices, feature_dict